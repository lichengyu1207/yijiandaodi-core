"""
合规治理层 Admin 配置

提供 AgentComplianceScore 的管理后台，支持将评分数据导出为 Excel：
- export_to_excel: 导出全部数据
- export_{safe,low,medium,high,critical}_to_excel: 按风险等级导出
"""

from django.contrib import admin
from django.http import HttpResponse

from .governance_models import AgentComplianceScore


@admin.register(AgentComplianceScore)
class AgentComplianceScoreAdmin(admin.ModelAdmin):
    """Agent合规性评分管理后台"""

    list_display = [
        'agent', 'overall_score', 'risk_level',
        'authentication_score', 'permission_score',
        'behavior_score', 'audit_score', 'updated_at'
    ]
    list_filter = ['risk_level']
    search_fields = ['agent__agent_id', 'agent__agent_name']
    readonly_fields = ['created_at', 'updated_at']

    # 风险等级显示名映射（与模型 choices 保持一致）
    RISK_LEVEL_LABELS = {
        'safe': '安全',
        'low': '低风险',
        'medium': '中风险',
        'high': '高风险',
        'critical': '严重风险',
    }

    EXPORT_HEADERS = [
        '序号', 'Agent ID', 'Agent名称', '综合评分', '风险等级',
        '认证评分', '权限评分', '行为评分', '审计评分',
        '违规次数', '近30天违规', '阻断操作次数',
        '最后操作时间', '24h操作次数', '7d操作次数', '30d操作次数',
        '评分更新时间', '创建时间'
    ]

    def has_change_permission(self, request, obj=None):
        """仅管理员/员工可变更（导出依赖此权限）"""
        if not request.user.is_authenticated:
            return False
        return request.user.is_superuser or request.user.is_staff

    # ==================== 导出方法 ====================

    def export_to_excel(self, request, queryset):
        """导出全部合规性评分数据为Excel"""
        try:
            from openpyxl import Workbook  # noqa: F401
        except ImportError:
            return None
        return self._generate_excel(queryset, None)

    def export_safe_to_excel(self, request, queryset):
        return self._export_by_risk_level(request, 'safe', self.RISK_LEVEL_LABELS['safe'])

    def export_low_to_excel(self, request, queryset):
        return self._export_by_risk_level(request, 'low', self.RISK_LEVEL_LABELS['low'])

    def export_medium_to_excel(self, request, queryset):
        return self._export_by_risk_level(request, 'medium', self.RISK_LEVEL_LABELS['medium'])

    def export_high_to_excel(self, request, queryset):
        return self._export_by_risk_level(request, 'high', self.RISK_LEVEL_LABELS['high'])

    def export_critical_to_excel(self, request, queryset):
        return self._export_by_risk_level(request, 'critical', self.RISK_LEVEL_LABELS['critical'])

    def _export_by_risk_level(self, request, risk_level, display_name):
        """按风险等级筛选并导出，数据为空时返回None"""
        queryset = self.model.objects.filter(risk_level=risk_level)
        if not queryset.exists():
            return None
        return self._generate_excel(queryset, display_name)

    @staticmethod
    def _to_excel_datetime(value):
        """openpyxl不支持带时区的datetime，转换为本地naive时间"""
        if value is None:
            return None
        from django.utils import timezone
        return timezone.localtime(value).replace(tzinfo=None)

    def _generate_excel(self, queryset, display_name):
        """生成Excel响应"""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from django.utils import timezone

        wb = Workbook()
        ws = wb.active
        ws.title = 'Agent合规性评分'

        # 标题行
        ws.append(self.EXPORT_HEADERS)

        # 标题样式（加粗、白字）
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, len(self.EXPORT_HEADERS) + 1):
            ws.cell(row=1, column=col).font = header_font

        # 数据行
        for idx, score in enumerate(queryset, start=1):
            ws.append([
                idx,
                score.agent.agent_id,
                score.agent.agent_name,
                score.overall_score,
                score.get_risk_level_display(),
                score.authentication_score,
                score.permission_score,
                score.behavior_score,
                score.audit_score,
                score.violations_count,
                score.violations_30d,
                score.blocked_operations_count,
                self._to_excel_datetime(score.last_operation_at),
                score.operations_24h,
                score.operations_7d,
                score.operations_30d,
                self._to_excel_datetime(score.score_updated_at),
                self._to_excel_datetime(score.created_at),
            ])

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 设置列宽
        for col in range(1, len(self.EXPORT_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # 文件名：Agent合规性评分_[风险等级]_YYYYMMDD_HHMMSS.xlsx
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Agent合规性评分_{timestamp}.xlsx'
        if display_name:
            filename = f'Agent合规性评分_{display_name}_{timestamp}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

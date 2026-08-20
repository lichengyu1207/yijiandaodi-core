"""
合规治理层导出功能单元测试

测试覆盖：
1. 不同风险等级的筛选逻辑
2. Excel文件生成
3. 数据完整性验证
4. 权限控制
5. 错误处理
"""

import os
import time
from io import BytesIO
from unittest.mock import Mock, patch
from django.test import TestCase, RequestFactory
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import load_workbook

from .governance_models import AgentComplianceScore
from .agent_identity_models import AgentIdentity
from .admin import AgentComplianceScoreAdmin

# 使用自定义用户模型
User = get_user_model()


def _content_disposition(response):
    """解码Content-Disposition头

    Django 对含非 ASCII 字符的响应头做 RFC 2047 编码，这里还原为原始字符串。
    """
    from email.header import decode_header

    parts = []
    for text, charset in decode_header(response['Content-Disposition']):
        if isinstance(text, bytes):
            parts.append(text.decode(charset or 'utf-8'))
        else:
            parts.append(text)
    return ''.join(parts)


class GovernanceExportTestCase(TestCase):
    """合规治理层导出功能测试基类"""
    
    def setUp(self):
        """测试准备"""
        # 创建测试用户
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        
        self.staff_user = User.objects.create_user(
            username='staff',
            email='staff@example.com',
            password='staff123',
            is_staff=True
        )
        
        self.normal_user = User.objects.create_user(
            username='user',
            email='user@example.com',
            password='user123'
        )
        
        # 创建测试Agent
        self.agents = {}
        risk_levels = ['safe', 'low', 'medium', 'high', 'critical']
        
        for i, risk_level in enumerate(risk_levels):
            agent = AgentIdentity.objects.create(
                agent_id=f'test_agent_{i+1}',
                agent_name=f'测试Agent{i+1}',
                trust_level=risk_level,
                api_key_hash=f'hash_{i+1}_{risk_level}'  # 提供唯一的api_key_hash
            )
            self.agents[risk_level] = agent
        
        # 创建合规性评分数据
        self.scores = {}
        score_values = {
            'safe': (95.0, 'safe', 100, 100, 95, 90),
            'low': (78.0, 'low', 85, 80, 75, 72),
            'medium': (65.0, 'medium', 70, 65, 60, 65),
            'high': (45.0, 'high', 50, 45, 40, 40),
            'critical': (25.0, 'critical', 30, 25, 20, 25)
        }
        
        for risk_level, (overall, level, auth, perm, behav, audit) in score_values.items():
            score = AgentComplianceScore.objects.create(
                agent=self.agents[risk_level],
                overall_score=overall,
                risk_level=level,
                authentication_score=auth,
                permission_score=perm,
                behavior_score=behav,
                audit_score=audit
            )
            self.scores[risk_level] = score
        
        # 创建请求工厂和管理后台实例
        self.factory = RequestFactory()
        self.admin = AgentComplianceScoreAdmin(AgentComplianceScore, None)
    
    def test_export_all_data(self):
        """测试导出全部数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 调用导出方法
        response = self.admin.export_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('Agent合规性评分_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证标题行
        self.assertEqual(ws.cell(row=1, column=1).value, '序号')
        self.assertEqual(ws.cell(row=1, column=5).value, '风险等级')
        
        # 验证数据行数（标题行 + 5条数据）
        self.assertEqual(ws.max_row, 6)
    
    def test_export_by_safe_level(self):
        """测试导出安全等级数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_safe_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertIn('_安全_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证只有安全等级的数据
        self.assertEqual(ws.max_row, 2)  # 标题行 + 1条数据
        self.assertEqual(ws.cell(row=2, column=5).value, '安全')
    
    def test_export_by_low_level(self):
        """测试导出低风险等级数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_low_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertIn('_低风险_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, '低风险')
    
    def test_export_by_medium_level(self):
        """测试导出中风险等级数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_medium_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertIn('_中风险_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, '中风险')
    
    def test_export_by_high_level(self):
        """测试导出高风险等级数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_high_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertIn('_高风险_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, '高风险')
    
    def test_export_by_critical_level(self):
        """测试导出严重风险等级数据"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_critical_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertIn('_严重风险_', _content_disposition(response))
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, '严重风险')
    
    def test_export_empty_queryset(self):
        """测试导出空数据集"""
        # 删除所有critical级别的数据
        AgentComplianceScore.objects.filter(risk_level='critical').delete()
        
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_critical_to_excel(
            request,
            AgentComplianceScore.objects.filter(risk_level='critical')
        )
        
        # 验证返回None（因为数据为空）
        self.assertIsNone(response)
    
    def test_export_data_integrity(self):
        """测试数据完整性"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_safe_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证所有字段都正确导出
        expected_headers = [
            '序号', 'Agent ID', 'Agent名称', '综合评分', '风险等级',
            '认证评分', '权限评分', '行为评分', '审计评分',
            '违规次数', '近30天违规', '阻断操作次数',
            '最后操作时间', '24h操作次数', '7d操作次数', '30d操作次数',
            '评分更新时间', '创建时间'
        ]
        
        for col_num, expected_header in enumerate(expected_headers, 1):
            self.assertEqual(
                ws.cell(row=1, column=col_num).value,
                expected_header,
                f"第{col_num}列标题不匹配"
            )
    
    def test_export_excel_formatting(self):
        """测试Excel格式化"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证标题行样式
        header_cell = ws.cell(row=1, column=1)
        self.assertTrue(header_cell.font.bold)
        self.assertEqual(header_cell.font.color.rgb, '00FFFFFF')
        
        # 验证冻结窗格
        self.assertEqual(ws.freeze_panes, 'A2')
        
        # 验证列宽已设置
        self.assertGreater(ws.column_dimensions['A'].width, 0)
        self.assertGreater(ws.column_dimensions['B'].width, 0)
    
    def test_export_permission_check(self):
        """测试导出权限检查"""
        # 测试普通用户（无权限）
        request = self.factory.post('/')
        request.user = self.normal_user
        
        # 验证权限检查
        self.assertFalse(
            self.admin.has_change_permission(request)
        )
        
        # 测试员工用户（有权限）
        request.user = self.staff_user
        self.assertTrue(
            self.admin.has_change_permission(request)
        )
    
    @patch('auth_app.admin.AgentComplianceScoreAdmin._export_by_risk_level')
    def test_export_method_calls(self, mock_export):
        """测试导出方法调用"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 测试每个导出方法都调用了通用方法
        self.admin.export_safe_to_excel(request, AgentComplianceScore.objects.all())
        mock_export.assert_called_with(request, 'safe', '安全')
        
        self.admin.export_low_to_excel(request, AgentComplianceScore.objects.all())
        mock_export.assert_called_with(request, 'low', '低风险')
        
        self.admin.export_medium_to_excel(request, AgentComplianceScore.objects.all())
        mock_export.assert_called_with(request, 'medium', '中风险')
        
        self.admin.export_high_to_excel(request, AgentComplianceScore.objects.all())
        mock_export.assert_called_with(request, 'high', '高风险')
        
        self.admin.export_critical_to_excel(request, AgentComplianceScore.objects.all())
        mock_export.assert_called_with(request, 'critical', '严重风险')
    
    def test_export_file_naming(self):
        """测试文件命名规范"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 测试不同风险等级的文件名
        response = self.admin.export_safe_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        self.assertRegex(
            _content_disposition(response),
            r'Agent合规性评分_安全_\d{8}_\d{6}\.xlsx'
        )
        
        response = self.admin.export_critical_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        self.assertRegex(
            _content_disposition(response),
            r'Agent合规性评分_严重风险_\d{8}_\d{6}\.xlsx'
        )
    
    def test_export_score_calculation(self):
        """测试评分数据导出的准确性"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_safe_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 获取数据库中的实际数据
        score = self.scores['safe']
        
        # 验证数据准确性
        self.assertEqual(float(ws.cell(row=2, column=4).value), score.overall_score)
        self.assertEqual(float(ws.cell(row=2, column=6).value), score.authentication_score)
        self.assertEqual(float(ws.cell(row=2, column=7).value), score.permission_score)
        self.assertEqual(float(ws.cell(row=2, column=8).value), score.behavior_score)
        self.assertEqual(float(ws.cell(row=2, column=9).value), score.audit_score)
    
    def test_export_with_violations(self):
        """测试带违规记录的数据导出"""
        # 为一个Agent添加违规记录
        score = self.scores['high']
        score.violations_count = 5
        score.violations_30d = 3
        score.blocked_operations_count = 2
        score.save()
        
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_high_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证违规数据导出
        self.assertEqual(ws.cell(row=2, column=10).value, 5)
        self.assertEqual(ws.cell(row=2, column=11).value, 3)
        self.assertEqual(ws.cell(row=2, column=12).value, 2)
    
    def test_export_multiple_agents_same_risk_level(self):
        """测试同一风险等级多个Agent的导出"""
        # 创建额外的安全等级Agent
        agent2 = AgentIdentity.objects.create(
            agent_id='test_agent_safe_2',
            agent_name='测试Agent安全2',
            trust_level='safe',
            api_key_hash='hash_safe_2'  # 提供唯一的api_key_hash
        )
        
        AgentComplianceScore.objects.create(
            agent=agent2,
            overall_score=92.0,
            risk_level='safe',
            authentication_score=95,
            permission_score=92,
            behavior_score=90,
            audit_score=91
        )
        
        request = self.factory.post('/')
        request.user = self.superuser
        
        response = self.admin.export_safe_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证两条安全等级数据
        self.assertEqual(ws.max_row, 3)  # 标题行 + 2条数据
        self.assertEqual(ws.cell(row=2, column=5).value, '安全')
        self.assertEqual(ws.cell(row=3, column=5).value, '安全')


class GovernanceExportPerformanceTestCase(TestCase):
    """导出功能性能测试"""
    
    def setUp(self):
        """创建大量测试数据"""
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        
        # 创建100个Agent
        for i in range(100):
            agent = AgentIdentity.objects.create(
                agent_id=f'perf_test_agent_{i}',
                agent_name=f'性能测试Agent{i}',
                trust_level='medium',
                api_key_hash=f'perf_hash_{i}'  # 提供唯一的api_key_hash
            )
            
            AgentComplianceScore.objects.create(
                agent=agent,
                overall_score=65.0,
                risk_level='medium',
                authentication_score=70,
                permission_score=65,
                behavior_score=60,
                audit_score=65
            )
        
        self.factory = RequestFactory()
        self.admin = AgentComplianceScoreAdmin(AgentComplianceScore, None)
    
    def test_export_large_dataset(self):
        """测试大数据集导出性能"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        start_time = time.time()
        response = self.admin.export_medium_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        elapsed_time = time.time() - start_time
        
        # 验证响应成功
        self.assertEqual(response.status_code, 200)
        
        # 验证性能（应该在5秒内完成）
        self.assertLess(elapsed_time, 5.0, "导出100条数据应该在5秒内完成")
        
        # 验证数据完整性
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 标题行 + 100条数据
        self.assertEqual(ws.max_row, 101)


class GovernanceExportErrorHandlingTestCase(TestCase):
    """导出功能错误处理测试"""
    
    def setUp(self):
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        
        self.factory = RequestFactory()
        self.admin = AgentComplianceScoreAdmin(AgentComplianceScore, None)
    
    def test_export_with_openpyxl_import_error(self):
        """测试openpyxl库导入失败的错误处理"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 模拟openpyxl导入失败
        with patch.dict('sys.modules', {'openpyxl': None}):
            response = self.admin.export_to_excel(
                request,
                AgentComplianceScore.objects.none()
            )
            
            # 应该返回None（错误被捕获）
            self.assertIsNone(response)
    
    def test_export_with_invalid_queryset(self):
        """测试无效查询集的处理"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 导出不存在的风险等级
        response = self.admin._export_by_risk_level(
            request,
            'nonexistent',
            '不存在'
        )
        
        # 应该返回None（数据为空）
        self.assertIsNone(response)


class GovernanceExportIntegrationTestCase(TestCase):
    """导出功能集成测试"""
    
    def setUp(self):
        """准备集成测试数据"""
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        
        # 创建一个完整的测试场景
        agent = AgentIdentity.objects.create(
            agent_id='integration_test_agent',
            agent_name='集成测试Agent',
            trust_level='medium',
            api_key_hash='integration_hash'  # 提供唯一的api_key_hash
        )
        
        self.score = AgentComplianceScore.objects.create(
            agent=agent,
            overall_score=75.5,
            risk_level='low',
            authentication_score=80,
            permission_score=75,
            behavior_score=72,
            audit_score=74,
            violations_count=1,
            violations_30d=0,
            blocked_operations_count=0,
            operations_24h=10,
            operations_7d=70,
            operations_30d=300
        )
        
        self.factory = RequestFactory()
        self.admin = AgentComplianceScoreAdmin(AgentComplianceScore, None)
    
    def test_full_export_workflow(self):
        """测试完整导出工作流"""
        request = self.factory.post('/')
        request.user = self.superuser
        
        # 执行导出
        response = self.admin.export_low_to_excel(
            request,
            AgentComplianceScore.objects.all()
        )
        
        # 验证响应
        self.assertEqual(response.status_code, 200)
        
        # 验证Excel内容
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # 验证所有数据字段
        self.assertEqual(ws.cell(row=2, column=2).value, 'integration_test_agent')
        self.assertEqual(ws.cell(row=2, column=3).value, '集成测试Agent')
        self.assertEqual(float(ws.cell(row=2, column=4).value), 75.5)
        self.assertEqual(ws.cell(row=2, column=5).value, '低风险')
        self.assertEqual(float(ws.cell(row=2, column=6).value), 80)
        self.assertEqual(float(ws.cell(row=2, column=7).value), 75)
        self.assertEqual(float(ws.cell(row=2, column=8).value), 72)
        self.assertEqual(float(ws.cell(row=2, column=9).value), 74)
        self.assertEqual(ws.cell(row=2, column=10).value, 1)
        self.assertEqual(ws.cell(row=2, column=11).value, 0)
        self.assertEqual(ws.cell(row=2, column=12).value, 0)
        self.assertEqual(ws.cell(row=2, column=14).value, 10)
        self.assertEqual(ws.cell(row=2, column=15).value, 70)
        self.assertEqual(ws.cell(row=2, column=16).value, 300)